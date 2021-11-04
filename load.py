import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from app.models import Information, ShipType, Date, VerifierCity


def parse_eedi(s):
    pass


workbook = load_workbook(r'C:\Users\Administrator\Documents\WeChat '
                         r'Files\wxid_c1iqy1sb51ea22\FileStorage\File\2021-11\huibt5110-main\Book3.xlsx')

sheet_name = workbook.sheetnames[0]
sheet = workbook[sheet_name]
gen = sheet.rows
for i in range(4):
    next(gen)

for row in gen:
    try:
        date = row[column_index_from_string('F') - 1].value
        date = datetime.datetime.strptime(date, '%d/%m/%Y').strftime('%Y-%m-%d')
        ship_type = row[column_index_from_string('C') - 1].value
        date = Date.objects.create(date=date)
        ship_type = ShipType.objects.get_or_create(name=ship_type)[0]
        verifier_city = row[column_index_from_string('K') - 1].value
        verifier_city = VerifierCity.objects.get_or_create(name=verifier_city)[0]

        imo = row[column_index_from_string('A') - 1].value
        eedi = row[column_index_from_string('E') - 1].value
        total_co2 = row[column_index_from_string('O') - 1].value
        total_time = row[column_index_from_string('D') - 1].value
        total_fuel = row[column_index_from_string('N') - 1].value

        # co2_emissions_per_distance = row[column_index_from_string('BC')]
        # co2_emissions_per_transport_work_mass = row(column_index_from_string('BD'))
        # co2_emissions_per_transport_work_volume = row(column_index_from_string('BE'))
        # co2_emissions_per_transport_work_dwt = row(column_index_from_string('BF'))
        # co2_emissions_per_transport_work_pax = row(column_index_from_string('BG'))
        # co2_emissions_per_transport_work_freight = row(column_index_from_string('BH'))
        Information.objects.create(
            imo=imo,
            eedi=eedi,
            total_co2=total_co2,
            total_time=total_time,
            total_fuel=total_fuel,
            date=date,
            ship_type=ship_type,
            verifier_city=verifier_city
            # co2_emissions_per_distance=co2_emissions_per_distance,
            # co2_emissions_per_transport_work_mass=co2_emissions_per_transport_work_mass,
            # co2_emissions_per_transport_work_volume=co2_emissions_per_transport_work_volume,
            # co2_emissions_per_transport_work_dwt=co2_emissions_per_transport_work_dwt,
            # co2_emissions_per_transport_work_pax=co2_emissions_per_transport_work_pax,
            # co2_emissions_per_transport_work_freight=co2_emissions_per_transport_work_freight
        )
    except Exception:
        print('错误')
